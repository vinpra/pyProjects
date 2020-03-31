# -*- coding: utf-8 -*-

import openpyxl as xl

wb = xl.load_workbook('ITEMS.xlsx', data_only=True)
sheet = wb['Rates']

for row in range(3,sheet.max_row +1):
    cell = sheet.cell(row,5)
    corected_price = cell.value * int(sheet.cell(row,6).value)
    corrected_price_cell = sheet.cell(row,7)
    corrected_price_cell.value = corected_price
    
wb.save('ITEMS.xlsx')

